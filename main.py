
import openpyxl
import os
from openpyxl.workbook import Workbook
import pandas as pd
from xlrd import open_workbook
import telebot
import shutil
import zipfile
import time




def main():


   def telegram_bot():
      bot = telebot.TeleBot("")


      @bot.message_handler(commands=['start'])
      def send_welcome(message):
         markup = telebot.types.ReplyKeyboardMarkup(row_width=2)
         itembtn1 = telebot.types.KeyboardButton('Обработай файл!')
         itembtn2 = telebot.types.KeyboardButton('По пивку?')
         itembtn3 = telebot.types.KeyboardButton('Пришли титьки!')
         markup.add(itembtn2, itembtn3, itembtn1)
         bot.send_message(message.chat.id, "Бонжур епта! Че надо?", reply_markup=markup)

         @bot.message_handler(func=lambda message: message.text == 'Обработай файл!')
         def option1(message):
            bot.send_message(message.chat.id, 'В п.зду работу!')
            time.sleep(1)
            bot.send_message(message.chat.id, 'Ладно, х.й с тобой. Кидай файл в чат')

         @bot.message_handler(func=lambda message: message.text == 'По пивку?')
         def option2(message):
            bot.send_message(message.chat.id, 'Это я за! ')
            bot.send_message(message.chat.id, 'https://i.pinimg.com/originals/b3/9c/0f/b39c0f73f6535fa3c5549c59cecc9ce3.gif')


         @bot.message_handler(func=lambda message: message.text == 'Пришли титьки!')
         def option3(message):
            bot.send_message(message.chat.id, 'https://www.slurmed.com/fgrabs/01acv10/01acv10_029.jpg')


      @bot.message_handler(content_types=['document'])
      def handle_docs_photo(message):
         bot.send_message(message.chat.id, 'Подожди. Пару сек...')
         file_id = message.document.file_id
         file_name = message.document.file_name
         file_info = bot.get_file(file_id)
         file = bot.download_file(file_info.file_path)
         try:
            with open(file_name, 'wb') as new_file:
               new_file.write(file)
         except:
            bot.send_message(message.chat.id, 'С файлом какаято фигня ')
            bot.send_message(message.chat.id, 'https://ananasposter.ru/image/catalog/poster/mult/87/6628.jpg')
            return

         # Get the file extension
         file_extension = file_name.split(".")[-1]
         # List of valid extensions
         valid_extensions = ['xls', 'xlsx']
         if file_extension not in valid_extensions:
            bot.send_message(message.chat.id, 'С файлом какаято фигня. Это вообще тот файл?')
            bot.send_message(message.chat.id, 'https://i.ytimg.com/vi/hENjphylnKM/maxresdefault.jpg')
            return


         # Distribute the file in folders
         folder_sent_file = file_name + "_folders"
         if not os.path.isdir(folder_sent_file):
            os.mkdir(folder_sent_file)
         # os.mkdir(file_name + "_folders")
         shutil.move(file_name, file_name + "_folders")
         os.chdir(file_name + "_folders")

         path_convert_file = file_name + "_folders/" + file_name
         path_file = file_name + "_folders"

         bot.send_message(message.chat.id, "https://giffiles.alphacoders.com/161/161758.gif")

         try:
            convert_file(file_name, path_file)
            # bot.send_message(message.chat.id, "За.ебись отконвертировалось")
         except:
            bot.send_message(message.chat.id, 'Произошла какая-то хуйня c конвертацией...')


         # bot.send_message(message.chat.id, "Распределяю по папкам")
         try:
            search_col_num()
            # bot.send_message(message.chat.id, "И этим без проблем")
         except:
            bot.send_message(message.chat.id, 'Произошла какая-то хуйня с поиском строки и обработкой таблиц...')


         # bot.send_message(message.chat.id, "Упаковываю")
         # Create a zip archive
         os.chdir("..")
         zip_name = file_name + ".zip"
         zip_file = zipfile.ZipFile(zip_name, 'w')
         for root, dirs, files in os.walk(path_file):
            for file in files:
               zip_file.write(os.path.join(root, file))
         zip_file.close()

         # Send the archive back to the user
         time.sleep(6)
         bot.send_document(message.chat.id, open(zip_name, 'rb'))
         time.sleep(2)


         bot.send_message(message.chat.id, "Данные по таблице удалены")
         # Clean up
         shutil.rmtree(file_name + "_folders")
         os.remove(zip_name)
         bot.send_message(message.chat.id, "С тебя пиво")
         bot.send_message(message.chat.id, "Нажми /start для выхода в меню")

      bot.polling()




   def convert_file(path_convert_file, path_file):
   # def convert_file(path_convert_file, path_file):
      wb = open_workbook(path_convert_file)
      print(wb)

      data = pd.read_excel(wb)

      data.to_excel('filename_new.xlsx', index=False)
      print("Converted")

   def search_col_num():
      # Open the Excel file
      wb = openpyxl.load_workbook('filename_new.xlsx')

      # Get the active sheet
      sheet = wb.active

      # Get the column name to search

      column_name_tab_1 = "Реестровый номер МО, куда направлен пациент"
      column_name_tab_6 = "Реестровый номер МО"

      # Iterate through the sheet and search for the column name
      for row in sheet.iter_rows():
         for cell in row:
            if cell.value == column_name_tab_1:
               # Print the column number
               print("Таблица 1. Нужный столбец:", cell.column)
               table_1()
            elif cell.value == column_name_tab_6:
               print("Таблица 6. Нужный столбец: ", cell.column)
               table_6()
               # return


   def table_1():
      path = 'filename_new.xlsx'
      number_of_col = 5
      # path = input("Введи полный путь до файла:")
      # number_of_col = int(input("Введи номер столбца по которому будет проходить фильтрация:"))



      # открывает файл экселя который мы потом сортируем
      wb_obj = openpyxl.load_workbook(path)
      sheet_obj = wb_obj.active
      row = sheet_obj.max_row
      columns = sheet_obj.max_column

      print(f'В файле: строк {row} и столбцов {columns}\n')
      # получаем список категорий (в данном случае) и упаковываем их в массив
      category_arr = []
      for i in range(4 , row+1):
         cell_obj = sheet_obj.cell( row=i, column = number_of_col)
         clear_value = cell_obj.value
         if not clear_value in category_arr:
            category_arr.append(clear_value)

      print(f"Список объектов {category_arr}")
      # создаем общую папку для всех проектов
      if not os.path.isdir("object"):
         os.mkdir("object")

      # обрабатываем таблицу и формируем папки
      for category in category_arr:

         sub_folder_name = str(category)
         print(f'Начало распределения объекта {sub_folder_name} ')
         # создаем папку для определенного проекта
         dir_project = "object/" + sub_folder_name
         if not os.path.isdir(dir_project):
            os.mkdir(dir_project)

         # заново перебираем ячейки с целью отфильтровать нужную категорию в файл

         # Создаем новый файл таблицы
         wb_category = Workbook()
         resultRow = 3
         wb_category_sheet = wb_category.active

         # Создаем первую строку с названием столбцов
         for i in range(1, columns+1):
            title_value = sheet_obj.cell(row=1, column= i).value
            wb_category_sheet.cell(row=1, column= i).value = title_value
            title_value = sheet_obj.cell(row=2, column=i).value
            wb_category_sheet.cell(row=2, column=i).value = title_value
         # Ищем название категории

         try:lookin_for = int(sub_folder_name)
         except:
            print(f'!!! Значение: {sub_folder_name}. Обработка цикла прекращена !!!')
            break
         for i in range(2, row):
            # Здесь мы определяемм значение ящейки в которой находится код
            value = sheet_obj.cell(row=i, column=number_of_col).value
            # print(value)
            try:value = int(value)
            except:pass
            # Сравниваем на совпаденение полученый код с тем что мы ищем в цикле
            if value == lookin_for:
               # Запускаем цикл в котором заполняем ячейки в новом файле одна за одной
               for j in range(1, columns + 1):
                  value = sheet_obj.cell(row=i, column=j).value
                  wb_category_sheet.cell(row=resultRow, column=j).value = value
                  # print(f'perezapisan {value}')
               resultRow += 1

         # print(f'Распределение объекта {lookin_for} завершено')

         # Сохраняем файл таблицы в определенную папку с определенным именем. Имя задается названием категории
         wb_category.save(dir_project + '/' + sub_folder_name + '_таблица_1.xlsx')
         print(f'   >>>Распределение объекта {lookin_for} завершена.\n   Файл сохранен: {sub_folder_name}.xlsx')
      print('Обработка данных завершена')


   def table_6():
      path = 'filename_new.xlsx'
      number_of_col = 2
      # path = input("Введи полный путь до файла:")
      # number_of_col = int(input("Введи номер столбца по которому будет проходить фильтрация:"))




      # открывает файл экселя который мы потом сортируем
      wb_obj = openpyxl.load_workbook(path)
      sheet_obj = wb_obj.active
      row = sheet_obj.max_row
      columns = sheet_obj.max_column

      print(f'В файле: строк {row} и столбцов {columns}\n')
      # получаем список категорий (в данном случае) и упаковываем их в массив
      category_arr = []
      for i in range(4 , row+1):
         cell_obj = sheet_obj.cell( row=i, column = number_of_col)
         clear_value = cell_obj.value
         if not clear_value in category_arr:
            category_arr.append(clear_value)

      print(f"Список объектов {category_arr}")
      # создаем общую папку для всех проектов
      if not os.path.isdir("object"):
         os.mkdir("object")

      # обрабатываем таблицу и формируем папки
      for category in category_arr:

         sub_folder_name = str(category)
         print(f'Начало распределения объекта {sub_folder_name} ')
         # создаем папку для определенного проекта
         sub_folder_name_dir = sub_folder_name.split(' ')[0]
         dir_project = "object/" + sub_folder_name_dir
         if not os.path.isdir(dir_project):
            os.mkdir(dir_project)

         # заново перебираем ячейки с целью отфильтровать нужную категорию в файл

         # Создаем новый файл таблицы
         wb_category = Workbook()
         resultRow = 4
         wb_category_sheet = wb_category.active

         # Создаем первую строку с названием столбцов
         for i in range(1, columns+1):
            title_value = sheet_obj.cell(row=1, column= i).value
            wb_category_sheet.cell(row=1, column= i).value = title_value
            title_value = sheet_obj.cell(row=2, column=i).value
            wb_category_sheet.cell(row=2, column=i).value = title_value
            title_value = sheet_obj.cell(row=3, column=i).value
            wb_category_sheet.cell(row=3, column=i).value = title_value
         # Ищем название категории

         lookin_for = sub_folder_name

         for i in range(4, row):
            # Здесь мы определяемм значение ящейки в которой находится код
            value = sheet_obj.cell(row=i, column=number_of_col).value

            # Сравниваем на совпаденение полученый код с тем что мы ищем в цикле
            if value == lookin_for:
               # Запускаем цикл в котором заполняем ячейки в новом файле одна за одной
               for j in range(1, columns + 1):
                  value = sheet_obj.cell(row=i, column=j).value
                  wb_category_sheet.cell(row=resultRow, column=j).value = value
               resultRow += 1


         # print(f'Распределение объекта {lookin_for} завершено')

         # Сохраняем файл таблицы в определенную папку с определенным именем. Имя задается названием категории
         wb_category.save(dir_project + '/' + sub_folder_name_dir + '_таблица_6.xlsx')
         print(f'   >>>Распределение объекта {lookin_for} завершена.\n   Файл сохранен: {sub_folder_name}.xlsx')
      print('Обработка данных завершена')


   def table_number():
      while True:
         table_number = int(input('Какую таблицу обрабатываем?'
                                  '\n если таблица номер 6 то введи цифру 6'
                                  '\n если таблица номер 1 то введи цифру 1: \n'))
         if table_number == 1:
            table_1()
            return True
         elif table_number == 6:
            table_6()
            return True
         else:
            print("Не понял. Давай еще раз")

   # Запускаем обработку функций здесь
   telegram_bot()
   # convert_file()
   # search_col_num()

   # не актуально
   # table_number()
   # table_1()
   # table_6()

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
   main()


